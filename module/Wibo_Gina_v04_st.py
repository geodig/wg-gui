# -*- coding: utf-8 -*-
"""
Created on Sat Sep 10 22:09:18 2022

@author: YOGB

WIBO_GINA (WITTEVEEN BOS GEOTECHNICAL INDONESIA)_version_0.4

"""

# IMPORT LIBRARIES ============================================================
import openpyxl
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.patches as patches
import matplotlib.text as mtext
import math
from scipy.interpolate import LinearNDInterpolator
import pandas as pd

# MISCELLANEOUS METHOD DEFINITION =============================================
class Estimation():
    def __init__(self,datax,datay,dataz):
        self.x = datax
        self.y = datay
        self.v = dataz

    def estimate(self,x,y,using='ISD'):
        """
        Estimate point at coordinate x,y based on the input data for this
        class.
        """
        if using == 'ISD':
            return self._isd(x,y)
        elif using == 'ISDrad':
            return self._isdradius(x,y)

    def _isd(self,x,y):
        d = np.sqrt((x-self.x)**2+(y-self.y)**2)
        if d.min() > 0:
            v = np.sum(self.v*(1/d**2)/np.sum(1/d**2))
            return v
        else:
            return self.v[d.argmin()]
    
    def _isdradius(self,x,y):
        radidw = 700
        d = np.sqrt((x-self.x)**2+(y-self.y)**2)
        d = d.flatten().tolist()
        
        xidw,yidw,vidw = [],[],[]
        for dis in d:
            if dis < radidw:
                indexdis = d.index(dis)
                xidw.append(self.x[indexdis])
                yidw.append(self.y[indexdis])
                vidw.append(self.v[indexdis])
        
        didw = np.sqrt((x-xidw)**2+(y-yidw)**2)
        if didw.min() > 0:
            v = np.sum(vidw*(1/didw**2)/np.sum(1/didw**2))
            return v
        else:
            return vidw[didw.argmin()]

def distance(x1,y1,x2,y2):
    D = np.sqrt((x1-x2)**2+(y1-y2)**2)
    return D

def read_shapefile(sf):
    """
    Read a shapefile into a Pandas dataframe with a 'coords' 
    column holding the geometry information. This uses the pyshp
    package
    """
    fields = [x[0] for x in sf.fields][1:]
    records = sf.records()
    shps = [s.points for s in sf.shapes()]
    df = pd.DataFrame(columns=fields, data=records)
    df = df.assign(coords=shps)
    return df

URL = r"http://a.tile.openstreetmap.org/{z}/{x}/{y}.png".format

TILE_SIZE = 256

def point_to_pixels(lon, lat, zoom):
    """convert gps coordinates to web mercator"""
    r = math.pow(2, zoom) * TILE_SIZE
    lat = math.radians(lat)
    x = int((lon + 180.0) / 360.0 * r)
    y = int((1.0 - math.log(math.tan(lat) + (1.0 / math.cos(lat))) / math.pi) / 2.0 * r)
    return x, y

def cpt_isbt(z, qc, fs, gwd):
    ndata = len(z)
     
    rf = [(i/j*100) for i,j in zip(fs,qc)]

    # unit weight (gamma) =========================================================
    gamma = [10*(0.27*np.log10(x)+0.36*np.log10(y/0.1)+1.236) for x,y in zip(rf,qc)]

    # vertical stress (sigma) =====================================================
    sig_tot = np.zeros(ndata)
    u0 = np.zeros(ndata)
    dz = z[1]-z[0]
    sig_tot[0] = z[0]*gamma[0]
    for i in range(ndata-1):
        sig_tot[i+1] = (sig_tot[i] + dz*gamma[i+1])
        if z[i+1] > gwd:
            u0[i+1] = (u0[i] + dz*10.0)
    sig_eff = sig_tot - u0

    # I-SBT (Robertson) ===========================================================
    Qt = [(x*1000-y)/z for x,y,z in zip(qc,sig_tot,sig_eff)]
    Fr = [a*1000/(b*1000-c)*100 for a,b,c in zip(fs,qc,sig_tot)]
    Ic = [np.sqrt((3.47-np.log10(x))**2 + (np.log10(y)+1.22)**2) for x,y in zip(Qt,Fr)]

    index = []                  # this is only for plotting purpose
    for i in range(ndata):
        if Ic[i] > 3.6:
            index.append(2)
        elif Ic[i] > 2.95 and Ic[i] <=3.6:
            index.append(3)
        elif Ic[i] > 2.60 and Ic[i] <=2.95:
            index.append(4)
        elif Ic[i] > 2.05 and Ic[i] <=2.60:
            index.append(5)
        elif Ic[i] > 1.31 and Ic[i] <=2.05:
            index.append(6)
        else:
            index.append(7)
    
    return(index)


# WIBO-GINA METHOD DEFINITION =================================================

class SoilModelCreator():
    def __init__(self, input_file, logwidth, spacewidth, valwidth, hvratio):
        self.filename = input_file
        self.logwidth = logwidth
        self.spacewidth = spacewidth
        self.valwidth = valwidth
        self.hvratio = hvratio

    def GetProfile(self, x_input, y_input, plotting_option='both', offset=10, surface='topo_data', generate_level=False):
        # READING INPUT DATA ==================================================

        wb = self.filename
        sheet_bhmgr = wb['borehole_manager']
        sheet_cptmgr = wb['cpt_manager']
        sheet_litholist = wb['lithology_list']
        sheet_lithology = wb['lithology']
        sheet_nspt = wb['nspt']
        sheet_cpt = wb['cpt']
        sheet_stratigraphy = wb['stratigraphy']
        sheet_topo = wb['topography']
        sheet_others = wb['others']
        
        plot_option = plotting_option  # lithology, stratigraphy, or both
        
        # BOREHOLE MANAGER, SURFACE, AND VERTICES DEFINITION===================
        
        elev_unit = sheet_others.cell(3,2).value
        
        BH_n = sheet_bhmgr.max_row - 1
        BH_ID,BH_X,BH_Y,BH_Z,BH_GWL = [],[],[],[],[]
        for i in range(BH_n):
            BH_ID.append(sheet_bhmgr.cell(i+2,2).value)
            BH_X.append(sheet_bhmgr.cell(i+2,3).value)
            BH_Y.append(sheet_bhmgr.cell(i+2,4).value)
            BH_Z.append(sheet_bhmgr.cell(i+2,5).value)
            BH_GWL.append(sheet_bhmgr.cell(i+2,6).value)
        
        BH_X = np.array(BH_X)
        BH_Y = np.array(BH_Y)
        BH_Z = np.array(BH_Z)
        BH_GWL = np.array(BH_GWL)
        
        CPT_n = sheet_cptmgr.max_row - 1
        CPT_ID, CPT_X, CPT_Y, CPT_Z, CPT_GWL = [],[],[],[],[]
        for i in range(CPT_n):
            CPT_ID.append(sheet_cptmgr.cell(i+2,2).value)
            CPT_X.append(sheet_cptmgr.cell(i+2,3).value)
            CPT_Y.append(sheet_cptmgr.cell(i+2,4).value)
            CPT_Z.append(sheet_cptmgr.cell(i+2,5).value)
            CPT_GWL.append(sheet_cptmgr.cell(i+2,6).value) 
        
        CPT_X = np.array(CPT_X)
        CPT_Y = np.array(CPT_Y)
        CPT_Z = np.array(CPT_Z)
        CPT_GWL = np.array(CPT_GWL)
        
        ALL_X = np.concatenate((BH_X, CPT_X))
        ALL_Y = np.concatenate((BH_Y, CPT_Y))
        ALL_Z = np.concatenate((BH_Z, CPT_Z))
        ALL_GWL = np.concatenate((BH_GWL, CPT_GWL))
        
        ALL_X = [i for i in ALL_X if i != None]
        ALL_Y = [i for i in ALL_Y if i != None]
        ALL_Z = [i for i in ALL_Z if i != None]
        ALL_GWL = [i for i in ALL_GWL if i != None]
        
        coord_x_max = np.max(ALL_X)
        coord_x_min = np.min(ALL_X)
        coord_y_max = np.max(ALL_Y)
        coord_y_min = np.min(ALL_Y)
        # coord_x_avg = np.average(ALL_X)
        # coord_y_avg = np.average(ALL_Y)
        coord_x_avg = (coord_x_max+coord_x_min)/2
        coord_y_avg = (coord_y_max+coord_y_min)/2
        coord_difx = coord_x_max - coord_x_min
        coord_dify = coord_y_max - coord_y_min
        coord_difmax = np.max([coord_difx, coord_dify])
        coeff = 0.7
        xmin = coord_x_avg - coeff*coord_difmax
        xmax = coord_x_avg + coeff*coord_difmax
        ymin = coord_y_avg - coeff*coord_difmax
        ymax = coord_y_avg + coeff*coord_difmax
        
        kamus_BH_X, kamus_BH_Y, kamus_BH_Z = {},{},{}
        for i in range(BH_n):
            kamus_BH_X[BH_ID[i]] = BH_X[i]
            kamus_BH_Y[BH_ID[i]] = BH_Y[i]
            kamus_BH_Z[BH_ID[i]] = BH_Z[i]
            
        kamus_CPT_X, kamus_CPT_Y, kamus_CPT_Z = {},{},{}
        for i in range(CPT_n):
            kamus_CPT_X[CPT_ID[i]] = CPT_X[i]
            kamus_CPT_Y[CPT_ID[i]] = CPT_Y[i]        
            kamus_CPT_Z[CPT_ID[i]] = CPT_Z[i]        
        
        
        # reading topography --------------------------------------------------
        topo_n = sheet_topo.max_row - 1
        topo_X, topo_Y, topo_Z = [],[],[]
        for i in range(topo_n):
            topo_X.append(sheet_topo.cell(i+2,1).value)
            topo_Y.append(sheet_topo.cell(i+2,2).value)
            topo_Z.append(sheet_topo.cell(i+2,3).value)
        
        topo_X = np.array(topo_X)
        topo_Y = np.array(topo_Y)
        topo_Z = np.array(topo_Z)        
        
        topo_X_all = np.concatenate((topo_X, BH_X, CPT_X))
        topo_Y_all = np.concatenate((topo_Y, BH_Y, CPT_Y))
        topo_Z_all = np.concatenate((topo_Z, BH_Z, CPT_Z))   
        
        
        topo_X_all = [i for i in topo_X_all if i != None]
        topo_Y_all = [i for i in topo_Y_all if i != None]
        topo_Z_all = [i for i in topo_Z_all if i != None]
        
        
        xmap = np.linspace(xmin, xmax, 200)
        ymap = np.linspace(ymin, ymax, 200)
        
        xmap1, ymap1 = np.meshgrid(xmap, ymap)
        
        f = LinearNDInterpolator(list(zip(topo_X_all,topo_Y_all)), topo_Z_all)
        
        zmap = np.zeros((len(xmap),len(ymap)))
        for i in range(len(xmap)):
            for j in range(len(ymap)):
                zmap [j][i] = f(xmap[i], ymap[j])
        
        
        xclick = x_input
        yclick = y_input
        
        jarak_poin_klik = []
        for i in range(len(xclick)-1):
            jarak_poin_klik.append(distance(xclick[i], yclick[i], xclick[i+1], yclick[i+1]))
        jarak_poin_klik = np.cumsum(jarak_poin_klik)
        
        vertices_poin_klik = np.concatenate(([0],jarak_poin_klik))
        
        ndiscrete = 100  
        radius = offset     
        
        x_discrete,y_discrete = [],[]
        for i in range(len(xclick)-1):
            listx,listy = [],[]
            for j in range(ndiscrete-1):
                listx.append(xclick[i]+(xclick[i+1]-xclick[i])/ndiscrete*(j+1))
                listy.append(yclick[i]+(yclick[i+1]-yclick[i])/ndiscrete*(j+1))
            x_discrete.append(listx)
            y_discrete.append(listy)
        
        x_discrete2, y_discrete2 = [],[]
        for i in range(len(xclick)-1):
            x_discrete2.append(np.concatenate((x_discrete[i],[xclick[i+1]])))
            y_discrete2.append(np.concatenate((y_discrete[i],[yclick[i+1]])))
        
        x_discrete2 = np.array(x_discrete2)
        y_discrete2 = np.array(y_discrete2)
        
        x_discrete3 = np.concatenate(([xclick[0]],x_discrete2.flatten()))
        y_discrete3 = np.concatenate(([yclick[0]],y_discrete2.flatten()))
        
        ground = Estimation(ALL_X,ALL_Y,ALL_Z)
        waterlevel = Estimation(ALL_X,ALL_Y,ALL_GWL)
        
        z_discrete,gwl_discrete = [],[]
        for i in range(len(x_discrete3)):
            if surface == 'topo_data':
                z_discrete.append(f(x_discrete3[i], y_discrete3[i]))
            elif surface == 'borehole_data':
                z_discrete.append(ground.estimate(x_discrete3[i], y_discrete3[i]))
            gwl_discrete.append(waterlevel.estimate(x_discrete3[i], y_discrete3[i]))
        
        jarak_vertices = []
        for i in range(len(x_discrete3)-1):
            jarak_vertices.append(distance(x_discrete3[i],y_discrete3[i],x_discrete3[i+1],y_discrete3[i+1]))
        
        jarak_vertices_cum = np.cumsum(jarak_vertices)
        jarak_vertices_cum0 = np.concatenate(([0],jarak_vertices_cum))
        
        # DEFINE WHICH BOREHOLE OR CPT IS INCLUDED WITHIN THE CROSS SECTION LINE
        # =====================================================================
        
        BH_included = []
        vertices_included = []
        for i in range(BH_n):
            jarak_tiap_bh = []
            for j in range(len(x_discrete3)):
                jarak_tiap_bh.append(distance(BH_X[i],BH_Y[i],x_discrete3[j],y_discrete3[j]))
            jarak_terdekat = np.min(jarak_tiap_bh)
            index_vertices = jarak_tiap_bh.index(jarak_terdekat)
            if jarak_terdekat <= radius:
                BH_included.append(BH_ID[i])
                vertices_included.append(jarak_vertices_cum0[index_vertices])
        
        CPT_included = []
        vertices_cpt_included = []
        for i in range(CPT_n):
            jarak_tiap_cpt = []
            for j in range(len(x_discrete3)):
                jarak_tiap_cpt.append(distance(CPT_X[i],CPT_Y[i],x_discrete3[j],y_discrete3[j]))
            jarak_terdekat = np.min(jarak_tiap_cpt)
            index_vertices = jarak_tiap_cpt.index(jarak_terdekat)
            if jarak_terdekat <= radius:
                CPT_included.append(CPT_ID[i])
                vertices_cpt_included.append(jarak_vertices_cum0[index_vertices])
        
        
        # DEFINE SCALING FOR NSPT AND CPT WIDTH ===============================
        
        item_included = BH_included + CPT_included
        x_item_included, y_item_included, z_item_included = [],[],[]
        for item in item_included:
            if item in BH_included:
                x_item_included.append(kamus_BH_X[item])
                y_item_included.append(kamus_BH_Y[item])
                z_item_included.append(kamus_BH_Z[item])
            elif item in CPT_included:
                x_item_included.append(kamus_CPT_X[item])
                y_item_included.append(kamus_CPT_Y[item])
                z_item_included.append(kamus_CPT_Z[item])
        
        jarakmin = []
        for i in range(len(item_included)):
            wadah = []
            for j in range(len(item_included)):
                jarak = distance(x_item_included[i],y_item_included[i],x_item_included[j],y_item_included[j])
                if jarak != 0.0:
                    wadah.append(jarak)
            jarakmin.append(min(wadah))
        
        shortestdist = np.average(jarakmin)
        
        distmax = np.max(jarak_vertices_cum0)
        coeff_width = self.valwidth
        coeff_scaling = shortestdist / distmax * coeff_width
        width_nspt = distmax*coeff_scaling
        ratio_nspt = width_nspt/50
        width_cpt = distmax*coeff_scaling
        ratio_cpt = width_cpt/30
        
        # NSPT ================================================================
        spt_n = int(sheet_nspt.max_column/2)
        spt_depth = np.zeros((spt_n,50))
        spt_value = np.zeros((spt_n,50))
        for i in range(spt_n):
            for j in range(50):
                spt_depth[i][j] = sheet_nspt.cell(j+3,2*i+1).value
                spt_value[i][j] = sheet_nspt.cell(j+3,2*i+2).value
        
        spt_depth2,spt_value2 = [],[]
        for i in range(spt_n):
            filtered_depth = spt_depth[i][~np.isnan(spt_depth[i])]
            filtered_nvalue = spt_value[i][~np.isnan(spt_value[i])]
            spt_depth2.append(filtered_depth)
            spt_value2.append(filtered_nvalue)
            
        spt_value_int = []
        for i in range(spt_n):
            spt_value_int.append(spt_value2[i].astype(int))
        
        spt_value_str = []
        for i in range(spt_n):
            spt_value_str.append(spt_value_int[i].astype(str))
        
        spt_depth_str = []
        for i in range(spt_n):
            spt_depth_str.append(spt_depth2[i].astype(str))
        
        spt_value_chart = [x*ratio_nspt for x in spt_value2]
        
        spt_elev = []
        for i in range(spt_n):
            filtered = []
            for j in range(len(spt_depth2[i])):
                filtered.append(BH_Z[i] - spt_depth2[i][j])
            spt_elev.append(filtered)
        
        # CPT =================================================================
        cpt_n = int(sheet_cpt.max_column/3)
        cpt_depth = np.zeros((cpt_n,2500))
        cpt_qc = np.zeros((cpt_n,2500))
        cpt_fs = np.zeros((cpt_n,2500))
        for i in range(cpt_n):
            for j in range(2500):
                cpt_depth[i][j] = sheet_cpt.cell(j+3,3*i+1).value
                cpt_qc[i][j] = sheet_cpt.cell(j+3,3*i+2).value
                cpt_fs[i][j] = sheet_cpt.cell(j+3,3*i+3).value
                
        cpt_depth2, cpt_qc2, cpt_fs2 = [],[],[]
        for i in range(cpt_n):
            filtered_depth = cpt_depth[i][~np.isnan(cpt_depth[i])]
            filtered_qc = cpt_qc[i][~np.isnan(cpt_qc[i])]
            filtered_fs = cpt_fs[i][~np.isnan(cpt_fs[i])]
            cpt_depth2.append(filtered_depth)
            cpt_qc2.append(filtered_qc)
            cpt_fs2.append(filtered_fs)
        
        cpt_fs3 = []
        for i in range(cpt_n):
            wadah = []
            for j in range(len(cpt_fs2[i])):
                wadah.append(np.abs(cpt_fs2[i][j]))
            cpt_fs3.append(wadah)
        
        cpt_qc3 = []
        for i in range(cpt_n):
            wadah = []
            for j in range(len(cpt_qc2[i])):
                wadah.append(np.abs(cpt_qc2[i][j]))
            cpt_qc3.append(wadah)
        
        isbt, gwd = [],[]
        for i in range(cpt_n):
            gwd.append(CPT_Z[i] - CPT_GWL[i])
            isbt.append(cpt_isbt(cpt_depth2[i], cpt_qc3[i], cpt_fs3[i], gwd[i]))

        cpt_qc_chart = [x*ratio_cpt for x in cpt_qc2]
        
        cpt_elev = []
        for i in range(cpt_n):
            filtered = []
            for j in range(len(cpt_depth2[i])):
                filtered.append(CPT_Z[i] - cpt_depth2[i][j])
            cpt_elev.append(filtered)

        # LITHOLOGY ===========================================================
        
        warna_litho,arsir_litho = {},{}
        nlithotype = sheet_litholist.max_row - 1
        for i in range(nlithotype):
            warna_litho[sheet_litholist.cell(2+i,1).value] = sheet_litholist.cell(2+i,2).value
            arsir_litho[sheet_litholist.cell(2+i,1).value] = sheet_litholist.cell(2+i,3).value
            
        litho_n = int(sheet_lithology.max_column/2)
        litho_depth = np.zeros((litho_n,50))
        
        litho_value = []
        for i in range(litho_n):
            filterer = []
            for j in range(50):
                litho_depth[i][j] = sheet_lithology.cell(j+3,2*i+1).value
                filterer.append(sheet_lithology.cell(j+3,2*i+2).value)
            litho_value.append(filterer)
        
        litho_depth2,litho_value2 = [],[]
        for i in range(litho_n):
            filtered_depth = litho_depth[i][~np.isnan(litho_depth[i])]
            filtered_value = [x for x in litho_value[i] if x != None]
            litho_depth2.append(filtered_depth)
            litho_value2.append(filtered_value)
        
        litho_depth3 = []
        for i in range(litho_n):
            litho_depth3.append(np.concatenate(([0],litho_depth2[i])))
        
        litho_elev = []
        for i in range(litho_n):
            filtered = []
            for j in range(len(litho_depth3[i])):
                filtered.append(BH_Z[i] - litho_depth3[i][j])
            litho_elev.append(filtered)
        
        # STRATIGRAPHY ========================================================
        
        warna_strati,arsir_strati = {},{}
        nstratitype = sheet_stratigraphy.max_row - 1
        for i in range(nstratitype):
            warna_strati[sheet_stratigraphy.cell(2+i,1).value] = sheet_stratigraphy.cell(2+i,2).value
            arsir_strati[sheet_stratigraphy.cell(2+i,1).value] = sheet_stratigraphy.cell(2+i,3).value
        
        strati_n = int(sheet_stratigraphy.max_column-3)
        
        strati_z = []
        for i in range(strati_n):
            header = sheet_stratigraphy.cell(1,4+i).value
            if header in BH_ID:
                strati_z.append(kamus_BH_Z[header])
            elif header in CPT_ID:
                strati_z.append(kamus_CPT_Z[header])
        
        strati_elev = []
        for i in range(nstratitype):
            elev = []
            for j in range(strati_n):
                x = sheet_stratigraphy.cell(2+i,4+j).value
                if x != None:
                    elev.append(strati_z[j]-x)
                else:
                    elev.append(None)
            strati_elev.append(elev)
        
              
        strati_BHID, strati_elev_filt = [],[]
        for i in range(nstratitype):
            bh,elev = [],[]
            for j in range(strati_n):
                if strati_elev[i][j] != None:
                    bh.append(sheet_stratigraphy.cell(1,4+j).value)
                    elev.append(strati_elev[i][j])
            strati_BHID.append(bh)
            strati_elev_filt.append(elev)
        
        strati_BHX, strati_BHY = [],[]
        for i in range(len(strati_BHID)):
            x,y = [],[]
            for j in range(len(strati_BHID[i])):
                if strati_BHID[i][j] in BH_ID:
                    x.append(kamus_BH_X[strati_BHID[i][j]])
                    y.append(kamus_BH_Y[strati_BHID[i][j]])
                elif strati_BHID[i][j] in CPT_ID:
                    x.append(kamus_CPT_X[strati_BHID[i][j]])
                    y.append(kamus_CPT_Y[strati_BHID[i][j]])
            strati_BHX.append(x)
            strati_BHY.append(y)
    
        def hitung(x,y):
            result = []
            for i in range(nstratitype):
                e = Estimation(strati_BHX[i],strati_BHY[i],strati_elev_filt[i])
                result.append(e.estimate(x,y))
            return result
        
        strati_interp = []
        for i in range(len(x_discrete3)):
            strati_interp.append(hitung(x_discrete3[i],y_discrete3[i]))
        strati_interp = np.array(strati_interp)
        
        strati_interp_all = []
        for i in range(len(x_discrete3)):
            strati_interp_all.append(np.concatenate(([z_discrete[i]], strati_interp[i])))
            
        strati_interp_all = np.array(strati_interp_all)
    
        strati_interp_all2 = []
        for i in range(len(x_discrete3)):
            wadah = []
            for j in range(nstratitype+1):
                if strati_interp_all[i][j] >= strati_interp_all[i][0]:
                    wadah.append(strati_interp_all[i][0])
                elif strati_interp_all[i][j] < strati_interp_all[i][0]:
                    wadah.append(strati_interp_all[i][j])
            strati_interp_all2.append(wadah)
        
        strati_interp_all2 = np.array(strati_interp_all2)
        strati_interp_trans2 = np.transpose(strati_interp_all2)
        
        strati_click, z_click = [],[]
        for i in range(len(xclick)):
            strati_click.append(hitung(xclick[i],yclick[i]))
            z_click.append(ground.estimate(xclick[i], yclick[i]))
        strati_click = np.array(strati_click)
        
        
        # ALGORITHM TO DEFINE WHICH LITHOLOGY, STRATIGRAPHY, 
        # AND NSPT TO BE INCLUDED IN THE CROSS SECTION ========================
        litho_elev_inc, litho_value2_inc, spt_elev_inc, spt_value_chart_inc, spt_value_str_inc = [],[],[],[],[]
        spt_depth_str_inc, BH_Z_inc = [],[]
        for i in range(len(BH_included)):
            index_inc = BH_ID.index(BH_included[i])
            litho_elev_inc.append(litho_elev[index_inc])
            litho_value2_inc.append(litho_value2[index_inc])
            spt_elev_inc.append(spt_elev[index_inc])
            spt_value_chart_inc.append(spt_value_chart[index_inc])
            spt_value_str_inc.append(spt_value_str[index_inc])
            spt_depth_str_inc.append(spt_depth_str[index_inc])
            BH_Z_inc.append(BH_Z[index_inc])

        cpt_elev_inc, cpt_qc_chart_inc, isbt_inc = [],[],[] 
        for i in range(len(CPT_included)):
            index_inc = CPT_ID.index(CPT_included[i])
            cpt_elev_inc.append(cpt_elev[index_inc])
            cpt_qc_chart_inc.append(cpt_qc_chart[index_inc])
            isbt_inc.append(isbt[index_inc])
        
        # PLOTTING ============================================================
        width_log = self.logwidth * distmax  
        space_log = self.spacewidth * distmax  
        
        cpt_qc_chart2 = []
        for i in range(len(CPT_included)):
            wadah = []
            for j in range(len(cpt_qc_chart_inc[i])):
                wadah.append(cpt_qc_chart_inc[i][j]+vertices_cpt_included[i]+width_log+space_log)
            cpt_qc_chart2.append(wadah)
        
        horz = vertices_included
        some_list,legend_list = [],[]
        litholist = list(warna_litho)
        stratilist = list(warna_strati)
        
        print(stratilist)
        
        elev_max = np.max(z_item_included)
        
        if any(litho_elev_inc) != bool([]):
            litho_elev_inc_flat = [j for sub in litho_elev_inc for j in sub]
            elev_min = np.min(litho_elev_inc_flat)
        else:
            cpt_elev_inc_flat = [j for sub in cpt_elev_inc for j in sub]
            elev_min = np.min(cpt_elev_inc_flat)
        
        class LegendTitle(object):
            def __init__(self, text_props=None):
                self.text_props = text_props or {}
                super(LegendTitle, self).__init__()
        
            def legend_artist(self, legend, orig_handle, fontsize, handlebox):
                x0, y0 = handlebox.xdescent, handlebox.ydescent
                title = mtext.Text(x0, y0, orig_handle,  **self.text_props)
                handlebox.add_artist(title)
                return title
        
        FIGSIZE = (14, 14/self.hvratio)
        fig,ax = plt.subplots(figsize=FIGSIZE)
        
        # PLOTTING GROUND SURFACE AND WATER LEVEL -----------------------------
        ax.plot(jarak_vertices_cum0,z_discrete,color='black',label='ground surface')
        ax.plot(jarak_vertices_cum0,gwl_discrete,color='blue',label='ground water')
        
        # PLOTTING CPT --------------------------------------------------------
        for i in range(len(CPT_included)):
            ax.plot(cpt_qc_chart2[i],cpt_elev_inc[i],color='navy')
            dz = cpt_elev_inc[i][0]-cpt_elev_inc[i][1]
            for j in range(len(cpt_elev_inc[i])-1):
                if isbt_inc[i][j+1] == 2:
                    colour = 'sienna'
                    label = 'peat'
                elif isbt_inc[i][j+1] == 3:
                    colour = 'olivedrab'
                    label = 'silty clay - clay'
                elif isbt_inc[i][j+1] == 4:
                    colour = 'yellowgreen'
                    label = 'clayey silt - silty clay'
                elif isbt_inc[i][j+1] == 5:
                    colour = 'lightsteelblue'
                    label = 'silty sand - sandy silt'
                elif isbt_inc[i][j+1] == 6:
                    colour = 'khaki'
                    label = 'clean sand - silty sand'
                elif isbt_inc[i][j+1] == 7:
                    colour = 'goldenrod'
                    label = 'gravel - dense sand'
                ax.add_patch(patches.Rectangle((vertices_cpt_included[i],cpt_elev_inc[i][j]),width_log,dz,facecolor=colour,zorder=3))
        
            xa = [vertices_cpt_included[i]+width_log,vertices_cpt_included[i]+width_log]
            xb = [vertices_cpt_included[i],vertices_cpt_included[i]]
            y = [kamus_CPT_Z[CPT_included[i]],np.min(cpt_elev_inc[i])]
            ax.plot(xa,y,c='k',zorder=3)
            ax.plot(xb,y,c='k',zorder=3)
            ax.plot([vertices_cpt_included[i],vertices_cpt_included[i]+width_log],[kamus_CPT_Z[CPT_included[i]],kamus_CPT_Z[CPT_included[i]]],c='k',zorder=3)
            ax.plot([vertices_cpt_included[i],vertices_cpt_included[i]+width_log],[np.min(cpt_elev_inc[i]),np.min(cpt_elev_inc[i])],c='k',zorder=3)
        
            x1 = [vertices_cpt_included[i],vertices_cpt_included[i]]
            y1 = [elev_min-5,elev_max+5]
            ax.plot(x1,y1,color='black',linestyle='--')
            ax.text(vertices_cpt_included[i],elev_max+5,CPT_included[i],rotation=45)
            
            ax.text(vertices_cpt_included[i]+width_log+0.5*width_cpt,kamus_CPT_Z[CPT_included[i]]+1.0,'qc (MPa)',horizontalalignment='center')
            
            ax.plot([vertices_cpt_included[i],vertices_cpt_included[i]+width_log+width_cpt],
                    [kamus_CPT_Z[CPT_included[i]],kamus_CPT_Z[CPT_included[i]]],c='grey',linewidth=0.8,linestyle='--')
            ax.plot([vertices_cpt_included[i],vertices_cpt_included[i]+width_log+width_cpt],
                    [np.min(cpt_elev_inc[i]),np.min(cpt_elev_inc[i])],c='grey',linewidth=0.8,linestyle='--')
            for j in range(6):
                ratio = (j+1)/6
                label = str((j+1)*5)
                xkotak = [vertices_cpt_included[i]+width_log+ratio*width_cpt,vertices_cpt_included[i]+width_log+ratio*width_cpt]
                ax.plot(xkotak,y,c='grey',linewidth=0.8,linestyle='--')
                ax.annotate(label,(xkotak[0]-1,kamus_CPT_Z[CPT_included[i]]+0.2),fontsize=9,horizontalalignment='center')
            
            cptdepth = kamus_CPT_Z[CPT_included[i]] - np.min(cpt_elev_inc[i])
            cptdepth = math.floor(cptdepth)
            cptdepth2 = np.arange(0,cptdepth+1,1)
            cptdepth_str = ['{:.1f}'.format(i) for i in cptdepth2]
            cptelev = [kamus_CPT_Z[CPT_included[i]] - x for x in cptdepth2]
            
            for j in range(len(cptelev)):
                ax.plot([vertices_cpt_included[i],vertices_cpt_included[i]-0.7*space_log],[cptelev[j],cptelev[j]],c='k')
                ax.annotate(cptdepth_str[j], (vertices_cpt_included[i]-1.2*space_log,cptelev[j]-0.15), horizontalalignment='right',fontsize=8)
            
            
        
        # PLOTTING STRATIGRAPHY -----------------------------------------------
        if plot_option == "stratigraphy" or plot_option == "both":
            ax.plot(jarak_vertices_cum0,strati_interp_all2,color='black',linewidth=0.8)
            for i in range(nstratitype):
                ax.fill_between(jarak_vertices_cum0,strati_interp_trans2[i],strati_interp_trans2[i+1],
                                color=warna_strati[stratilist[i]],ec='gray',
                                hatch=arsir_strati[stratilist[i]],label=stratilist[i],alpha=0.3,zorder=1)
        
        # PLOTTING NSPT PROFILE -----------------------------------------------
        for i in range(len(BH_included)):
            if plot_option == "lithology" or plot_option == "both":
                xnspt = horz[i]+width_log+space_log
            else:
                xnspt = horz[i]
            
            ax.barh(spt_elev_inc[i],spt_value_chart_inc[i],left=xnspt,color='navy',edgecolor='black',zorder=3)
            for j in range(len(spt_value_str_inc[i])):
                ax.annotate(spt_value_str_inc[i][j], (spt_value_chart_inc[i][j]+xnspt+space_log,spt_elev_inc[i][j]-0.5), color='black')
                ax.annotate(spt_depth_str_inc[i][j], (horz[i]-1.0*space_log,spt_elev_inc[i][j]-0.25), horizontalalignment = 'right',fontsize=8)
                ax.plot([horz[i],horz[i]-0.7*space_log],[spt_elev_inc[i][j],spt_elev_inc[i][j]],color='black')
           
            x = [xnspt,xnspt]
            y = [BH_Z_inc[i],np.min(spt_elev_inc[i])]
            ax.plot(x,y,c='k')
            ax.plot([x[0],x[0]+50*ratio_nspt],[BH_Z_inc[i],BH_Z_inc[i]],c='k')
            ax.text(x[0]+0.5,BH_Z_inc[i]+1,'N-SPT value')
            
            x1 = [horz[i],horz[i]]
            y1 = [elev_min-5,elev_max+5]
            ax.plot(x1,y1,color='black',linestyle='--')
            ax.text(horz[i],elev_max+5,BH_included[i],rotation=45)
        
        # PLOTTING VERTICES LINE AND LABEL ------------------------------------
        for i in range(len(vertices_poin_klik)):
            x2 = [vertices_poin_klik[i],vertices_poin_klik[i]]
            y2 = [elev_min-5,elev_max+5]
            ax.plot(x2,y2,color='silver',linestyle='-.',zorder=1)
            # ax.text(vertices_poin_klik[i],elev_max+5,'vert %d'%(i+1),rotation=45,color='silver')
        
        # PLOTTING LITHOLOGY PROFILE AND LEGEND -------------------------------
        if len(CPT_included) != 0:
            warnacpt = ['sienna','olivedrab','yellowgreen','lightsteelblue','khaki','goldenrod']
            labelcpt = ['peat','silty clay - clay','clayey silt - silty clay','silty sand - sandy silt','clean sand - silty sand','gravel - dense sand']
            for i in range(6):
                ax.add_patch(patches.Rectangle((0,-1000),1,1,facecolor=warnacpt[i],label=labelcpt[i]))
                
        
        
        if plot_option == "lithology" or plot_option == "both":
            for i in range(len(BH_included)):
                for j in range(len(litho_value2_inc[i])):
                    ax.add_patch(patches.Rectangle((horz[i],litho_elev_inc[i][j]), 
                                                    width_log, litho_elev_inc[i][j+1]-litho_elev_inc[i][j], 
                                                    # label=litho_value2[i][j] if litho_value2[i][j] not in some_list else '', 
                                                    color=warna_litho[litho_value2_inc[i][j]], ec='black', 
                                                    hatch=arsir_litho[litho_value2_inc[i][j]],zorder=3))
            
                    if litho_value2_inc[i][j] not in some_list:
                        legend_list.append(litho_value2_inc[i][j])
                    
                    some_list.append(litho_value2_inc[i][j])
        
        legend_index = []    
        for i in range(len(legend_list)):
            legend_index.append(litholist.index(legend_list[i]))
        
        legend_index_sorted = sorted(legend_index)
        
        legend_index2 = []
        for i in range(len(legend_list)):
            legend_index2.append(legend_index_sorted.index(legend_index[i]))
            
        legend_list_sorted = []
        for i in range(len(legend_list)):
            indexs = legend_index2.index(i)
            legend_list_sorted.append(legend_list[indexs])
        
        for i in range(len(legend_list)):
            ax.add_patch(patches.Rectangle((0,-200),1,1,color=warna_litho[legend_list_sorted[i]],
                                            hatch=arsir_litho[legend_list_sorted[i]],ec='black',
                                            label=legend_list_sorted[i]))
        
        a = nstratitype
        h,l = ax.get_legend_handles_labels()
        
        if len(CPT_included) == 0 and plot_option == "both":
            ax.legend(['Legend'] + h[:2] + ['','Stratigraphy'] + h[2:a+2] + ['','Lithology'] + h[a+2:], 
                      ['']       + l[:2] + ['','']             + l[2:a+2] + ['','']          + l[a+2:],
                        handler_map={str: LegendTitle({'fontsize': 13})},
                      bbox_to_anchor=(1, 1),edgecolor='none')
        elif len(CPT_included) == 0 and plot_option == "stratigraphy":
            ax.legend(['Legend'] + h[:2] + ['','Stratigraphy'] + h[2:], 
                      ['']       + l[:2] + ['','']             + l[2:],
                        handler_map={str: LegendTitle({'fontsize': 13})},
                      bbox_to_anchor=(1, 1),edgecolor='none')
        elif len(CPT_included) == 0 and plot_option == "lithology":
            ax.legend(['Legend'] + h[:2] + ['','Lithology']    + h[2:], 
                      ['']       + l[:2] + ['','']             + l[2:],
                        handler_map={str: LegendTitle({'fontsize': 13})},
                      bbox_to_anchor=(1, 1),edgecolor='none')
        elif len(CPT_included) != 0 and plot_option == "both":
            ax.legend(['Legend'] + h[:2] + ['','Stratigraphy'] + h[2:a+2] + ['','Lithology'] + h[a+8:] + ['','CPT Robertson (2010)'] + h[2+a:8+a], 
                      ['']       + l[:2] + ['','']             + l[2:a+2] + ['','']          + l[a+8:] + ['','']                     + l[2+a:8+a],
                        handler_map={str: LegendTitle({'fontsize': 13})},
                      bbox_to_anchor=(1, 1),edgecolor='none')
        elif len(CPT_included) != 0 and plot_option == "stratigraphy":
            ax.legend(['Legend'] + h[:2] + ['','Stratigraphy'] + h[2:2+a] + ['','CPT Robertson (2010)'] + h[2+a:], 
                      ['']       + l[:2] + ['','']             + l[2:2+a] + ['','']                     + l[2+a:],
                        handler_map={str: LegendTitle({'fontsize': 13})},
                      bbox_to_anchor=(1, 1),edgecolor='none')
        elif len(CPT_included) != 0 and plot_option == "lithology":
            ax.legend(['Legend'] + h[:2] + ['','Lithology']    + h[8:] + ['','CPT Robertson (2010)'] + h[2:8], 
                      ['']       + l[:2] + ['','']             + l[8:] + ['','']                     + l[2:8],
                        handler_map={str: LegendTitle({'fontsize': 13})},
                      bbox_to_anchor=(1, 1),edgecolor='none')
        
        # OTHER PLOT SETTING --------------------------------------------------
        
        ax.set_ylim(elev_min-5,elev_max+10)
        ax.grid(linewidth=0.3,color='moccasin')
        ax.set_xlabel('horizontal distance (m)')
        ax.set_ylabel('elevation (%s)'%(elev_unit))
        plt.tight_layout()
        
        # WRITING AND SAVING TEXT FILE ========================================
        if generate_level == True:
            outputname = 'generated_level.xlsx'
            wb = openpyxl.Workbook()
            wb.remove(worksheet=wb['Sheet'])
            sh = wb.create_sheet('level')
            sh.cell(1,1).value = 'layer_name'
            for i in range(len(stratilist)):
                sh.cell(2*i+2,1).value = stratilist[i] + '_top'
                sh.cell(2*i+3,1).value = stratilist[i] + '_bot'
            for i in range(len(xclick)):
                for j in range(len(stratilist)):
                    sh.cell(2*j+3,i+2).value = strati_click[i][j]
                for k in range(len(stratilist)-1):
                    sh.cell(2*k+4,i+2).value = strati_click[i][k]
                sh.cell(2,i+2).value = z_click[i]
                sh.cell(1,i+2).value = 'point_%d'%(i+1)
            wb.save(filename=outputname)
        
        return(fig)
        
        # END OF SCRIPT =======================================================

