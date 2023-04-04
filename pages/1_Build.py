# -*- coding: utf-8 -*-
"""
Created on Thu Mar 30 09:59:45 2023

@author: YOGB
"""

# IMPORT LIBRARIES ============================================================
import streamlit as st
import folium
from streamlit_folium import st_folium
from folium.plugins import Draw
from matplotlib import pyplot as plt
import pandas as pd
import openpyxl
import utm
import numpy as np
from scipy.interpolate import griddata
from openpyxl import Workbook
from module.Wibo_Gina_v04_st import SoilModelCreator
import geojsoncontour
import geojson
from zipfile import ZipFile
import base64
from streamlit_authenticator import Authenticate
import yaml
from yaml import SafeLoader
from streamlit_extras.app_logo import add_logo

# INITIAL AND DEFAULT SETTING =================================================
st.set_page_config(page_title="WIBOGINA", layout="wide")
add_logo("./logo1_small.png", height=70)
listlevel = ["Borehole level"]
filename = "input_wibogina.xlsx"
zipObj = ZipFile("input_file_compiled.zip", "w")

# AUTHENTICATION ==============================================================
with open('./config.yml') as file:
    config = yaml.load(file, Loader=SafeLoader)

authenticator = Authenticate(
    config['credentials'],
    config['cookie']['name'],
    config['cookie']['key'],
    config['cookie']['expiry_days'],
    config['preauthorized']
)

if st.session_state["authentication_status"]:
    authenticator.logout('Logout', 'sidebar')
    st.sidebar.header(f'Welcome *{st.session_state["name"]}*')

# CONTENT =====================================================================
    tab1, tab2 = st.tabs(["Map","Section Plot"])

    with st.sidebar.expander("**INPUT FILES**"):
        uploaded_files = st.file_uploader("Upload BH/CPT files:", type=["xlsx"], accept_multiple_files=True)
        uploaded_topo = st.file_uploader("Upload topography data:", type=["xlsx"], accept_multiple_files=False) 
        ncontour = st.slider("Number of lines for contour:", min_value=5, max_value=50, step=1)
        colcontour = st.color_picker("Choose color for contour lines:")
        uploaded_shp = st.file_uploader("Upload GeoJSON file:", type=["geojson"], accept_multiple_files=True)
    submit_file = st.sidebar.button("UPLOAD AND COMPILE", use_container_width=True)
    
       
    # READING AND DISPLAY ALL INPUT FILES =========================================
    # AND CREATION OF WIBOGINA INPUT FILE =========================================
    if submit_file:
    # INITIATE WIBOGINA INPUT FILE ------------------------------------------------
        wb_out = Workbook()
        wb_out.remove(worksheet=wb_out['Sheet'])
        
        sheet_bhmgr = wb_out.create_sheet('borehole_manager')
        sheet_cptmgr = wb_out.create_sheet('cpt_manager')
        sheet_litho = wb_out.create_sheet('lithology')
        sheet_nspt = wb_out.create_sheet('nspt')
        sheet_cpt = wb_out.create_sheet('cpt')
        sheet_litholist = wb_out.create_sheet('lithology_list')
        sheet_strati = wb_out.create_sheet('stratigraphy')
        sheet_topo = wb_out.create_sheet('topography')
        sheet_other = wb_out.create_sheet('others')
        
    # BOREHOLE OR CPT FILES -------------------------------------------------------
    # DISPLAY ---------------------------------------------------------------------
        if uploaded_files:
            listfile = [i.name for i in uploaded_files]
        
            ID, BH_label, X, Y, Z, water_level, elev_unit, UTM_zone, UTM_NS, city, province, project_code = [],[],[],[],[],[],[],[],[],[],[],[]
            lithnaming, surveyor, date, remark, max_depth = [],[],[],[],[]
            for file in uploaded_files:
                wb = openpyxl.load_workbook(file, data_only=True)
                sheet = wb["general"]
                ID.append(sheet.cell(1,2).value)
                BH_label.append(sheet.cell(2,2).value)
                X.append(sheet.cell(3,2).value)
                Y.append(sheet.cell(4,2).value)
                Z.append(sheet.cell(5,2).value)
                water_level.append(sheet.cell(6,2).value)
                elev_unit.append(sheet.cell(7,2).value)
                UTM_zone.append(sheet.cell(8,2).value)
                UTM_NS.append(sheet.cell(9,2).value)
                city.append(sheet.cell(10,2).value)
                province.append(sheet.cell(11,2).value)
                project_code.append(sheet.cell(12,2).value)
                lithnaming.append(sheet.cell(13,2).value)
                surveyor.append(sheet.cell(14,2).value)
                remark.append(sheet.cell(16,2).value)
                if isinstance(sheet.cell(15,2).value,str):
                    date.append(sheet.cell(15,2).value)
                elif isinstance(sheet.cell(15,2).value,int):
                    date.append(str(sheet.cell(15,2).value))
                else:
                    date.append(str(sheet.cell(15,2).value))
                
                if "spt" in file.name:
                    sheet2 = wb['nspt']
                    spt_n = sheet2.max_row - 1
                    spt_depth = []
                    for i in range(spt_n):
                        spt_depth.append(sheet2.cell(i+2,1).value)
                    bottom = np.max(spt_depth)
                    max_depth.append(bottom)
                elif "cpt" in file.name:
                    sheet2 = wb['cpt']
                    spt_n = sheet2.max_row - 1
                    spt_depth = []
                    for i in range(spt_n):
                        spt_depth.append(sheet2.cell(i+2,1).value)
                    bottom = np.max(spt_depth)
                    max_depth.append(bottom)
            
            data = {'ID':ID,
                   'BH_label':BH_label,
                   'X_UTM':X,
                   'Y_UTM':Y,
                   'Z':Z,
                   'water_level':water_level,
                   'elev_unit':elev_unit,
                   'UTM_zone': UTM_zone,
                   'UTM_NS': UTM_NS,
                   'city':city,
                   'provice':province,
                   'project_code':project_code,
                   'naming':lithnaming,
                   'surveyor':surveyor,
                   'date':date,
                   'remark':remark,
                   'max_depth':max_depth}
            
            df = pd.DataFrame(data)
            
            topo_zone = df["UTM_zone"].iloc[0]
            topo_hemi = df["UTM_NS"].iloc[0]
            
            latlon, lat, lon = [],[],[]
            for i in range(len(df)):
                if df["UTM_NS"].iloc[i] == 'N':
                    hemis = True
                elif df["UTM_NS"].iloc[i] == 'S':
                    hemis = False
                latlon.append(utm.to_latlon(df["X_UTM"].iloc[i], df["Y_UTM"].iloc[i], df["UTM_zone"].iloc[i], northern=hemis))
                lat.append(latlon[i][0])
                lon.append(latlon[i][1])
            
            df['lat']=lat
            df['lon']=lon
            
            latmean = df['lat'].mean()
            lonmean = df['lon'].mean()
            location = [latmean, lonmean]
            m = folium.Map(location=location, zoom_start=16
                            # , tiles="CartoDB positron"
                            )
            
            icons = []
            for i in range(0,len(df)):
                if "spt" in df['ID'].iloc[i]:
                    icons.append(folium.Icon(color='black'))
                elif "cpt" in df['ID'].iloc[i]:
                    icons.append(folium.Icon(color='red'))
            
            for i in range(0,len(df)):
                folium.Marker([df['lat'].iloc[i], df['lon'].iloc[i]], popup="ID: %s\nLabel: %s"%(df['ID'].iloc[i],df['BH_label'].iloc[i]), icon=icons[i]).add_to(m)
            
            draw = Draw(export=False,
                        filename="shapefile.geojson",
                        position='topleft',
                        draw_options={'polyline': {'allowIntersection': False,
                                                   'shapeOptions': {'stroke': True,
                                                         			'color': '#DC0303',
                                                         			'weight': 4,
                                                         			'opacity': 0.7}},
                                      'polygon': False,
                                      'rectangle': False,
                                      'circle': False,
                                      'marker': False,
                                      'circlemarker': False},
                        edit_options={'poly': {'allowIntersection': False}}) 
            
            draw.add_to(m)
    
    # COMPILE ---------------------------------------------------------------------
            field_bhmgr = ['No','BH_ID','X','Y','Z','GWL']
            field_cptmgr = ['No','CPT_ID','X','Y','Z','GWL']
            for i in range(len(field_bhmgr)):
                sheet_bhmgr.cell(1,i+1).value = field_bhmgr[i]
                sheet_cptmgr.cell(1,i+1).value = field_cptmgr[i]
                
            for i in range(len(df)):
                if 'spt' in df['ID'].iloc[i]:
                    sheet_bhmgr.cell(i+2,1).value = i+1
                    sheet_bhmgr.cell(i+2,2).value = df['BH_label'].iloc[i]
                    sheet_bhmgr.cell(i+2,3).value = df['X_UTM'].iloc[i]
                    sheet_bhmgr.cell(i+2,4).value = df['Y_UTM'].iloc[i]
                    sheet_bhmgr.cell(i+2,5).value = df['Z'].iloc[i]
                    sheet_bhmgr.cell(i+2,6).value = df['water_level'].iloc[i]
                    sheet_bhmgr.cell(i+2,7).value = df['elev_unit'].iloc[i]
                if 'cpt' in df['ID'].iloc[i]:
                    sheet_cptmgr.cell(i+2,1).value = i+1
                    sheet_cptmgr.cell(i+2,2).value = df['BH_label'].iloc[i]
                    sheet_cptmgr.cell(i+2,3).value = df['X_UTM'].iloc[i]
                    sheet_cptmgr.cell(i+2,4).value = df['Y_UTM'].iloc[i]
                    sheet_cptmgr.cell(i+2,5).value = df['Z'].iloc[i]
                    sheet_cptmgr.cell(i+2,6).value = df['water_level'].iloc[i]
                    sheet_cptmgr.cell(i+2,7).value = df['elev_unit'].iloc[i]
            
            sheet_other.cell(1,1).value = 'shapefile_path'
            sheet_other.cell(2,1).value = 'shapefile_name'
            sheet_other.cell(3,1).value = 'elevation_unit'
            elevunit1 = [i for i in df['elev_unit']]
            elevunit2 = []
            for i in elevunit1:
                if i not in elevunit2:
                    elevunit2.append(i)
            if len(elevunit2) > 1:
                sheet_other.cell(3,2).value = 'non-uniform elevation unit'
            elif len(elevunit2) == 1:
                sheet_other.cell(3,2).value = df['elev_unit'].iloc[0]
            
            for i in range(len(df)):
                if 'spt' in df['ID'].iloc[i]:
                    sheet_litho.cell(1,2*i+1).value = df['BH_label'].iloc[i]
                    sheet_litho.cell(2,2*i+1).value = 'depth'
                    sheet_litho.cell(2,2*i+2).value = 'lithology'
                    sheet_nspt.cell(1,2*i+1).value = df['BH_label'].iloc[i]
                    sheet_nspt.cell(2,2*i+1).value = 'depth'
                    sheet_nspt.cell(2,2*i+2).value = 'nspt'
                    wb_in = openpyxl.load_workbook(uploaded_files[i], data_only=True)
                    sh_lith = wb_in['lithology']
                    sh_nspt = wb_in['nspt']
                    sh_lithtype = wb_in['litholist']
                    for j in range(sh_lith.max_row - 1):
                        sheet_litho.cell(j+3,2*i+1).value = sh_lith.cell(j+2,1).value
                        sheet_litho.cell(j+3,2*i+2).value = sh_lith.cell(j+2,2).value
                    for k in range(sh_nspt.max_row - 1):
                        sheet_nspt.cell(k+3,2*i+1).value = sh_nspt.cell(k+2,1).value
                        sheet_nspt.cell(k+3,2*i+2).value = sh_nspt.cell(k+2,2).value
                    for x in range(sh_lithtype.max_row):
                        sheet_litholist.cell(x+1,1).value = sh_lithtype.cell(x+1,1).value
                        sheet_litholist.cell(x+1,2).value = sh_lithtype.cell(x+1,2).value            
                        sheet_litholist.cell(x+1,3).value = sh_lithtype.cell(x+1,3).value            
                if 'cpt' in df['ID'].iloc[i]:
                    sheet_cpt.cell(1,3*i+1).value = df['BH_label'].iloc[i]
                    sheet_cpt.cell(2,3*i+1).value = 'depth'
                    sheet_cpt.cell(2,3*i+2).value = 'qc'
                    sheet_cpt.cell(2,3*i+3).value = 'fs'
                    wb_in = openpyxl.load_workbook(uploaded_files[i], data_only=True)
                    sh_cpt = wb_in['cpt']
                    for j in range(sh_cpt.max_row - 1):
                        sheet_cpt.cell(j+3,3*i+1).value = sh_cpt.cell(j+2,1).value
                        sheet_cpt.cell(j+3,3*i+2).value = sh_cpt.cell(j+2,2).value
                        sheet_cpt.cell(j+3,3*i+3).value = sh_cpt.cell(j+2,3).value
    
    # TOPOGRAPHY/BATHYMETRY -------------------------------------------------------
    # DISPLAY ---------------------------------------------------------------------
        if uploaded_shp:
            for file in uploaded_shp:
                gj = geojson.load(file)
                folium.GeoJson(gj, name=file.name).add_to(m)
            if uploaded_topo:
                pass
            else:
                folium.LayerControl().add_to(m)

        if uploaded_topo:
            listlevel.append("Topo/bathy data")
            wbtopo = openpyxl.load_workbook(uploaded_topo, data_only=True)
            shtopo = wbtopo[wbtopo.sheetnames[0]]
            
            if topo_hemi == "N":
                hemi = True
            elif topo_hemi == "S":
                hemi = False
            
            topox, topoy, topoz, topolat, topolon = [],[],[],[],[]
            for i in range(shtopo.max_row-1):
                topox.append(shtopo.cell(i+2,1).value)
                topoy.append(shtopo.cell(i+2,2).value)
                topoz.append(shtopo.cell(i+2,3).value)
                sheet_topo.cell(i+2,1).value = topox[i]
                sheet_topo.cell(i+2,2).value = topoy[i]
                sheet_topo.cell(i+2,3).value = topoz[i]
                topolatlon = utm.to_latlon(topox[i], topoy[i], zone_number=topo_zone, northern=hemi)
                topolat.append(topolatlon[0])
                topolon.append(topolatlon[1])
        
            interval = np.arange(np.min(topoz),np.max(topoz),(np.max(topoz)-np.min(topoz))/ncontour).tolist()
        
            y = np.linspace(min(topolat), max(topolat), 1000)
            x = np.linspace(min(topolon), max(topolon), 1000)
            
            X,Y = np.meshgrid(x, y)
            Z = griddata((topolon, topolat), topoz, (X, Y), method='cubic')
        
            contour = plt.contour(X,Y,Z, levels=interval)
            kontur = geojsoncontour.contour_to_geojson(contour=contour,
                                                        ndigits=5)
            
            style = {'color':colcontour,
                     'weight':1.5}
            
            folium.GeoJson(kontur, name="Topo/bathy contour", style_function=lambda x:style).add_to(m)
            folium.LayerControl().add_to(m)
            st.session_state["listlevel"] = listlevel
            
        st.session_state["label"] = BH_label
        
        st.session_state["map"] = m
        st.session_state["wb_out"] = wb_out
    
        wb_out.save(filename=filename)
    
        zipObj.write(filename)
        zipObj.close()
        ZipfileDotZip = "input_file_compiled.zip"
        
        with open(ZipfileDotZip, "rb") as f:
            bytes = f.read()
            b64 = base64.b64encode(bytes).decode()
            href = f"<a href=\"data:file/zip;base64,{b64}\" download='{ZipfileDotZip}.zip'>\
                DOWNLOAD COMPILED INPUT FILE\
            </a>"
            st.session_state["href"]=href
    
    # DISPLAY FOLIUM MAP INSIDE TAB1 --------------------------------------
    if "map" in st.session_state:
        m = st.session_state["map"]
        with tab1:
            mymap = st_folium(m, height=700, width=1200
                                , returned_objects=["last_active_drawing"]
                                )
    # READING CLICKED COORDINATES FROM POLYLINE -----------------------------------
        if mymap["last_active_drawing"]:
            
            with st.sidebar.expander("**CHART SETTING**"):
                hvratio = st.slider("Horizontal/vertical plot ratio:", value=2.0, min_value=1.0, max_value=3.0, step=0.1)
                valwidth = st.slider("CPT/SPT width:", value=0.5, min_value=0.1, max_value=0.9, step=0.1)
                logwidth = st.slider("Lithology log width:", value=0.02, min_value=0.01, max_value=0.05, step=0.005)
                spacewidth = st.slider("In-between space width:", value=0.004, min_value=0.001, max_value=0.01, step=0.001)
                if "listlevel" in st.session_state:
                    listlevel = st.session_state["listlevel"]
                elif "listlevel" not in st.session_state:
                    listlevel = listlevel
                plot_options = ["lithology","stratigraphy","both"]
                surface = st.selectbox("Ground surface line:", options=listlevel)
                plottype = st.selectbox("Plotting options:", options=plot_options)
                offset = st.slider("Offset distance from centerline:", value=10, min_value=5, max_value=50, step=1)
            draw_section = st.sidebar.button("DRAW SECTION", use_container_width=True)
            
            features = mymap["last_active_drawing"]["geometry"]["coordinates"]
            UTM, X_UTM, Y_UTM = [],[],[]
            for i in range(len(features)):
                UTM.append(utm.from_latlon(features[i][1], features[i][0]))
                X_UTM.append(UTM[i][0])
                Y_UTM.append(UTM[i][1])
            st.session_state["xutm"] = X_UTM
            st.session_state["yutm"] = Y_UTM
                
    # CALLING WIBOGINA ============================================================
            if draw_section:
                if "wb_out2" in st.session_state:
                    wb = st.session_state["wb_out2"]
                elif "wb_out2" not in st.session_state:
                    wb = st.session_state["wb_out"]
                
                project = SoilModelCreator(wb, logwidth, spacewidth, valwidth, hvratio)
                
                if surface == "Borehole level":
                    sur = "borehole_data"
                elif surface == "Topo/bathy data":
                    sur = "topo_data"
                
                X_UTM = st.session_state["xutm"]
                Y_UTM = st.session_state["yutm"]
                
                fig = project.GetProfile(x_input=X_UTM, y_input=Y_UTM, plotting_option=plottype, offset=offset, surface=sur)
                st.session_state["fig"] = fig
                
        with tab2:
            if "fig" in st.session_state:
                fig = st.session_state["fig"]
                st.pyplot(fig)
                
            elif "fig" not in st.session_state:
                st.warning("Define the section path by drawing a polyline in the map.",icon="⚠️")
            
            with st.expander("STRATIGRAPHY INPUT"):
                if "label" in st.session_state:
                    BH_label = st.session_state["label"]
                    st.header("Stratigraphy")
                    st.markdown("""
                                This is where you define the stratigraphy, section by section. You can add or delete the row of the table
                                to increase/decrease your stratigraphy unit. For the color, you can use the color picker below the table 
                                and copy-paste the HEX color code. You will need to input the bottom depth of your stratigraphy into each
                                borehole/CPT column in the table.
                                """)
                    
                    emptylist = [None]*3
                    emptylist2 = [5,10,15]
                    strati_dict = {'stratigraphy':["A","B","C"],
                                    'color':["#39778c","#276a86","#1E5A7E"],
                                    'hatch':emptylist}
                    for i in range(len(BH_label)):
                        strati_dict[BH_label[i]]=emptylist2
                    
                    strati_df = pd.DataFrame(strati_dict)
                    strati_table = st.experimental_data_editor(strati_df, use_container_width=True, num_rows="dynamic")
                    st.color_picker("Use this color picker to determine the HEX color code for your stratigraphy unit:")
                    st.session_state["strati_table"]=strati_table
                    update_strati = st.button("UPDATE STRATIGRAPHY")
                    
                    if update_strati:
                        wb_out2 = st.session_state["wb_out"]
                        sh_strati = wb_out2["stratigraphy"]
                        
                        strati_table = st.session_state["strati_table"]
                        strati_header = strati_table.columns.values.tolist()
                        strati_nrows = len(strati_table)+1
                        strati_ncols = len(strati_header)
                        
                        for i in range(strati_ncols):
                            sh_strati.cell(1,i+1).value = strati_header[i]
                        
                        for j in range(strati_nrows-1):
                            sh_strati.cell(j+2,1).value = strati_table["stratigraphy"].iloc[j]
                            sh_strati.cell(j+2,2).value = strati_table["color"].iloc[j]
                            sh_strati.cell(j+2,3).value = strati_table["hatch"].iloc[j]
                            
                        for i in range(strati_ncols-3):
                            for j in range(strati_nrows-1):
                                if strati_table[strati_header[i+3]].iloc[j] != None:
                                    sh_strati.cell(j+2,i+4).value = float(strati_table[strati_header[i+3]].iloc[j])
                                elif strati_table[strati_header[i+3]].iloc[j] == None:
                                    sh_strati.cell(j+2,i+4).value = strati_table[strati_header[i+3]].iloc[j]
                        
                        st.session_state["wb_out2"] = wb_out2
                        
                        wb_out2.save(filename=filename)
    
                        zipObj.write(filename)
                        zipObj.close()
                        ZipfileDotZip = "input_file_compiled.zip"
                        
                        with open(ZipfileDotZip, "rb") as f:
                            bytes = f.read()
                            b64 = base64.b64encode(bytes).decode()
                            href2 = f"<a href=\"data:file/zip;base64,{b64}\" download='{ZipfileDotZip}.zip'>\
                                DOWNLOAD COMPILED INPUT FILE\
                            </a>"
                            st.session_state["href2"]=href2
                        
                elif "label" not in st.session_state:
                    st.warning("Please upload input files before filling this table.",icon="⚠️")
    
        if "href2" in st.session_state:
            href = st.session_state["href2"]
            st.markdown(href, unsafe_allow_html=True)
        elif "href2" not in st.session_state and "href" in st.session_state:
            href = st.session_state["href"]
            st.markdown(href, unsafe_allow_html=True)
        elif "href2" not in st.session_state and "href" not in st.session_state:
            pass
                        
        
    elif "map" not in st.session_state:
        tab1.warning("Upload input files and click 'Upload'.",icon="⚠️")
        tab2.warning("Upload input files, click 'Upload', draw polyline, and click 'Draw Section'.",icon="⚠️")

# END OF CONTENT ==============================================================

elif st.session_state["authentication_status"] == False:
    st.error('Username/password is incorrect')
elif st.session_state["authentication_status"] == None:
    st.warning('Please enter your username and password from the home page')



