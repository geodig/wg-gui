# -*- coding: utf-8 -*-
"""
Created on Mon Apr  3 13:31:05 2023

@author: YOGB
"""
# IMPORT LIBRARIES ============================================================
import streamlit as st
import folium
from streamlit_folium import st_folium
from folium.plugins import Draw
from matplotlib import pyplot as plt
import pandas as pd
import openpyxl
import utm
import numpy as np
from scipy.interpolate import griddata
from module.Wibo_Gina_v04_st import SoilModelCreator
import geojsoncontour
import geojson
from zipfile import ZipFile
import streamlit_authenticator as stauth
import yaml
from yaml import SafeLoader
from streamlit_extras.app_logo import add_logo

# INITIAL AND DEFAULT SETTING =================================================
st.set_page_config(page_title="WIBOGINA", layout="wide")
add_logo("./logo1_small.png", height=70)
listlevel = ["Borehole level"]
filename = "input_wibogina.xlsx"
zipObj = ZipFile("input_file_compiled.zip", "w")

# AUTHENTICATION ==============================================================
with open('./config.yml') as file:
    config = yaml.load(file, Loader=SafeLoader)

authenticator = stauth.Authenticate(
    config['credentials'],
    config['cookie']['name'],
    config['cookie']['key'],
    config['cookie']['expiry_days'],
    config['preauthorized']
)

if st.session_state["authentication_status"]:
    authenticator.logout('Logout', 'sidebar')
    st.sidebar.header(f'Welcome *{st.session_state["name"]}*')

# CONTENT =====================================================================
    tab1, tab2 = st.tabs(["Map","Section Plot"])
            
    with st.sidebar.expander("**INPUT FILES**"):
        uploaded_file = st.file_uploader("Upload WIBOGINA input file:", type=["xlsx"], accept_multiple_files=False)
        utm_zone = st.number_input("UTM zone:", min_value=46, max_value=54)
        utm_hemi = st.selectbox("UTM hemisphere:", options=["North","South"])
        ncontour = st.slider("Number of lines for contour:", min_value=5, max_value=50, step=1)
        colcontour = st.color_picker("Choose color for contour lines:")
        uploaded_shp = st.file_uploader("Upload GeoJSON file:", type=["geojson"], accept_multiple_files=True)
    submit_file = st.sidebar.button("UPLOAD", use_container_width=True)
    
    if submit_file:
        if uploaded_file:
            wb_in = openpyxl.load_workbook(uploaded_file, data_only=True)
            st.session_state["wb_in"] = wb_in
            
            # READING INPUT SPREADSHEET ===================================================
            sheet_bhmgr = wb_in['borehole_manager']
            sheet_cptmgr = wb_in['cpt_manager']
            sheet_topo = wb_in["topography"]
    
            BH_n = sheet_bhmgr.max_row - 1
            BH_ID,BH_X,BH_Y,BH_Z,BH_GWL,BH_cat = [],[],[],[],[],[]
            for i in range(BH_n):
                BH_ID.append(sheet_bhmgr.cell(i+2,2).value)
                BH_X.append(sheet_bhmgr.cell(i+2,3).value)
                BH_Y.append(sheet_bhmgr.cell(i+2,4).value)
                BH_Z.append(sheet_bhmgr.cell(i+2,5).value)
                BH_GWL.append(sheet_bhmgr.cell(i+2,6).value)
                BH_cat.append("BH")
    
            BH_X = np.array(BH_X)
            BH_Y = np.array(BH_Y)
            BH_Z = np.array(BH_Z)
            BH_GWL = np.array(BH_GWL)
            BH_ID = np.array(BH_ID)
            BH_cat = np.array(BH_cat)
    
            CPT_n = sheet_cptmgr.max_row - 1
            CPT_ID, CPT_X, CPT_Y, CPT_Z, CPT_GWL, CPT_cat = [],[],[],[],[],[]
            for i in range(CPT_n):
                CPT_ID.append(sheet_cptmgr.cell(i+2,2).value)
                CPT_X.append(sheet_cptmgr.cell(i+2,3).value)
                CPT_Y.append(sheet_cptmgr.cell(i+2,4).value)
                CPT_Z.append(sheet_cptmgr.cell(i+2,5).value)
                CPT_GWL.append(sheet_cptmgr.cell(i+2,6).value) 
                CPT_cat.append("CPT")
    
            CPT_X = np.array(CPT_X)
            CPT_Y = np.array(CPT_Y)
            CPT_Z = np.array(CPT_Z)
            CPT_GWL = np.array(CPT_GWL)
            CPT_ID = np.array(CPT_ID)
            CPT_cat = np.array(CPT_cat)
    
            ALL_X = np.concatenate((BH_X, CPT_X))
            ALL_Y = np.concatenate((BH_Y, CPT_Y))
            ALL_Z = np.concatenate((BH_Z, CPT_Z))
            ALL_GWL = np.concatenate((BH_GWL, CPT_GWL))
            ALL_ID = np.concatenate((BH_ID, CPT_ID))
            ALL_cat = np.concatenate((BH_cat, CPT_cat))
    
            ALL_X = [i for i in ALL_X if i != None]
            ALL_Y = [i for i in ALL_Y if i != None]
            ALL_Z = [i for i in ALL_Z if i != None]
            ALL_GWL = [i for i in ALL_GWL if i != None]
            ALL_ID = [i for i in ALL_ID if i != None]
            ALL_cat = [i for i in ALL_cat if i != None]
    
            data = {'ID' : ALL_ID,
                    'X_UTM' : ALL_X,
                    'Y_UTM' : ALL_Y,
                    'Z' : ALL_Z,
                    'GWL' : ALL_GWL,
                    'CLASS' : ALL_cat}
    
            df = pd.DataFrame(data)
    
            latlon, lat, lon = [],[],[]
            for i in range(len(ALL_X)):
                if utm_hemi == 'North':
                    hemis = True
                elif utm_hemi == 'South':
                    hemis = False
                latlon.append(utm.to_latlon(ALL_X[i], ALL_Y[i], utm_zone, northern=hemis))
                lat.append(latlon[i][0])
                lon.append(latlon[i][1])
    
            df['lat']=lat
            df['lon']=lon
    
            latmean = df['lat'].mean()
            lonmean = df['lon'].mean()
            location = [latmean, lonmean]
            m = folium.Map(location=location, zoom_start=16
                            # , tiles="CartoDB positron"
                           )
    
            icons = []
            for i in range(0,len(df)):
                if df['CLASS'].iloc[i] == "BH":
                    icons.append(folium.Icon(color='black'))
                elif df['CLASS'].iloc[i] == "CPT":
                    icons.append(folium.Icon(color='red'))
    
            for i in range(0,len(df)):
                folium.Marker([df['lat'].iloc[i], df['lon'].iloc[i]], popup=df['ID'].iloc[i], icon=icons[i]).add_to(m)
    
            draw = Draw(export=False,
                        filename="shapefile.geojson",
                        position='topleft',
                        draw_options={'polyline': {'allowIntersection': False,
                                                   'shapeOptions': {'stroke': True,
                                                         			'color': '#DC0303',
                                                         			'weight': 4,
                                                         			'opacity': 0.7}},
                                      'polygon': False,
                                      'rectangle': False,
                                      'circle': False,
                                      'marker': False,
                                      'circlemarker': False},
                        edit_options={'poly': {'allowIntersection': False}}) 
    
            draw.add_to(m)
            
            if uploaded_shp:
                for file in uploaded_shp:
                    gj = geojson.load(file)
                    folium.GeoJson(gj, name=file.name).add_to(m)
                if sheet_topo.max_row > 10:
                    pass
                else:
                    folium.LayerControl().add_to(m)
            
            if sheet_topo.max_row > 10:
                listlevel.append("Topo/bathy data")
                
                if utm_hemi == "North":
                    hemi = True
                elif utm_hemi == "South":
                    hemi = False
                
                topox, topoy, topoz, topolat, topolon = [],[],[],[],[]
                for i in range(sheet_topo.max_row-1):
                    topox.append(sheet_topo.cell(i+2,1).value)
                    topoy.append(sheet_topo.cell(i+2,2).value)
                    topoz.append(sheet_topo.cell(i+2,3).value)
                    topolatlon = utm.to_latlon(topox[i], topoy[i], zone_number=utm_zone, northern=hemi)
                    topolat.append(topolatlon[0])
                    topolon.append(topolatlon[1])
            
                interval = np.arange(np.min(topoz),np.max(topoz),(np.max(topoz)-np.min(topoz))/ncontour).tolist()
            
                y = np.linspace(min(topolat), max(topolat), 1000)
                x = np.linspace(min(topolon), max(topolon), 1000)
                
                X,Y = np.meshgrid(x, y)
                Z = griddata((topolon, topolat), topoz, (X, Y), method='cubic')
            
                contour = plt.contour(X,Y,Z, levels=interval)
                kontur = geojsoncontour.contour_to_geojson(contour=contour,
                                                            ndigits=5)
                
                style = {'color':colcontour,
                         'weight':1.5}
                
                folium.GeoJson(kontur, name="Topo/bathy contour", style_function=lambda x:style).add_to(m)
                folium.LayerControl().add_to(m)
                st.session_state["listlevel_2"] = listlevel
            
            st.session_state["label_2"] = ALL_ID
            
            st.session_state["map_2"] = m
    
    # DISPLAY FOLIUM MAP INSIDE TAB1 --------------------------------------
    if "map_2" in st.session_state:
        m = st.session_state["map_2"]
        with tab1:
            mymap = st_folium(m, height=700, width=1200
                                , returned_objects=["last_active_drawing"]
                                )
        
    # READING CLICKED COORDINATES FROM POLYLINE -----------------------------------
        if mymap["last_active_drawing"]:
            
            with st.sidebar.expander("**CHART SETTING**"):
                hvratio = st.slider("Horizontal/vertical plot ratio:", value=2.0, min_value=1.0, max_value=3.0, step=0.1)
                valwidth = st.slider("CPT/SPT width:", value=0.5, min_value=0.1, max_value=0.9, step=0.1)
                logwidth = st.slider("Lithology log width:", value=0.02, min_value=0.01, max_value=0.05, step=0.005)
                spacewidth = st.slider("In-between space width:", value=0.004, min_value=0.001, max_value=0.01, step=0.001)
                if "listlevel_2" in st.session_state:
                    listlevel = st.session_state["listlevel_2"]
                elif "listlevel_2" not in st.session_state:
                    listlevel = listlevel
                plot_options = ["lithology","stratigraphy","both"]
                surface = st.selectbox("Ground surface line:", options=listlevel)
                plottype = st.selectbox("Plotting options:", options=plot_options)
                offset = st.slider("Offset distance from centerline:", value=10, min_value=5, max_value=50, step=1)
            draw_section = st.sidebar.button("DRAW SECTION", use_container_width=True)
            
            features = mymap["last_active_drawing"]["geometry"]["coordinates"]
            UTM, X_UTM, Y_UTM = [],[],[]
            for i in range(len(features)):
                UTM.append(utm.from_latlon(features[i][1], features[i][0]))
                X_UTM.append(UTM[i][0])
                Y_UTM.append(UTM[i][1])
            st.session_state["xutm_2"] = X_UTM
            st.session_state["yutm_2"] = Y_UTM
                
    # CALLING WIBOGINA ============================================================
            if draw_section:
                wb = st.session_state["wb_in"]
                
                project = SoilModelCreator(wb, logwidth, spacewidth, valwidth, hvratio)
                
                if surface == "Borehole level":
                    sur = "borehole_data"
                elif surface == "Topo/bathy data":
                    sur = "topo_data"
                
                X_UTM = st.session_state["xutm_2"]
                Y_UTM = st.session_state["yutm_2"]
                
                fig = project.GetProfile(x_input=X_UTM, y_input=Y_UTM, plotting_option=plottype, offset=offset, surface=sur)
                st.session_state["fig_2"] = fig
                
        with tab2:
            if "fig_2" in st.session_state:
                fig = st.session_state["fig_2"]
                st.pyplot(fig)
                
            elif "fig_2" not in st.session_state:
                st.warning("Define the section path by drawing a polyline in the map.",icon="⚠️")
        
    
    elif "map_2" not in st.session_state:
        tab1.warning("Upload input files and click 'Upload'.",icon="⚠️")
        tab2.warning("Upload input files, click 'Upload', draw polyline, and click 'Draw Section'.",icon="⚠️")
    

# END OF CONTENT ==============================================================

elif st.session_state["authentication_status"] == False:
    st.error('Username/password is incorrect')
elif st.session_state["authentication_status"] == None:
    st.warning('Please enter your username and password from the home page')

